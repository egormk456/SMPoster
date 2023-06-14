import asyncio

from bot import dp
from data.commands import setter
from data import db_gino


async def add_users():
    await db_gino.on_startup(dp)
    from openpyxl import load_workbook

    book = load_workbook(filename=f"users.xlsx")
    sheet = book["users"]

    for row in range(2, sheet.max_row + 1):
        if sheet["A" + str(row)].value is not None:
            try:
                await setter.client_add(sheet["A" + str(row)].value,
                                        sheet["B" + str(row)].value)
            except:
                pass


if __name__ == '__main__':
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    loop.run_until_complete(add_users())
